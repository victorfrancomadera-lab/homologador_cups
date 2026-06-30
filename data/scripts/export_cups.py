"""
Exporta la Tabla de Referencia CUPS (SISPRO/Minsalud) a CSV normalizado.

Entrada: TablaReferencia_CUPS__1.xlsx (hoja 'Table')
  Columnas relevantes: Codigo, Nombre, Descripcion (capitulo), Habilitado (SI/NO)
Salida: data/output/cups.csv
  codigo, nombre, capitulo, vigente (bool)
"""
import csv
from pathlib import Path

import openpyxl

XLSX = Path(r"C:\Users\vfran\Downloads\TablaReferencia_CUPS__1.xlsx")
OUT = Path(__file__).resolve().parents[1] / "output" / "cups.csv"
OUT.parent.mkdir(parents=True, exist_ok=True)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    vigentes = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        codigo = r[idx["Codigo"]]
        if codigo is None:
            continue
        codigo = str(codigo).strip()
        nombre = (r[idx["Nombre"]] or "").strip()
        capitulo = (r[idx["Descripcion"]] or "").strip()
        hab = str(r[idx["Habilitado"]] or "").strip().upper()
        vigente = hab in ("SI", "SÍ", "S", "TRUE", "1")
        vigentes += vigente
        rows.append({
            "codigo": codigo,
            "nombre": nombre,
            "capitulo": capitulo,
            "vigente": "1" if vigente else "0",
        })

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        wr = csv.DictWriter(f, fieldnames=["codigo", "nombre", "capitulo", "vigente"])
        wr.writeheader()
        wr.writerows(rows)

    print(f"CUPS total: {len(rows)} | vigentes: {vigentes} | no vigentes: {len(rows)-vigentes}")
    print(f"Salida: {OUT}")


if __name__ == "__main__":
    main()
